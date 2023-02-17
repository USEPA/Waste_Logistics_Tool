import arcpy,os,sys;
import json;
import openpyxl;
from openpyxl import Workbook;
from openpyxl.styles import Font,Alignment;

import source.util;
import source.obj_AllHazardsWasteLogisticsTool;
import source.obj_Scenario;
import source.obj_QuerySet;
import source.obj_Waste;

###############################################################################
import importlib
importlib.reload(source.util);
importlib.reload(source.obj_AllHazardsWasteLogisticsTool);
importlib.reload(source.obj_Scenario);
importlib.reload(source.obj_QuerySet);
importlib.reload(source.obj_Waste);

const_version = "2.0";

def execute(self, parameters, messages):

   #########################################################################
   # Step 10
   # Abend if edits are pending
   #########################################################################
   if source.util.sniff_editing_state():
      raise arcpy.ExecuteError("Error.  Pending edits must be saved or cleared before proceeding.");
      
   waste = source.obj_Waste.Waste();
      
   #########################################################################
   # Step 20
   # Read the parameters
   #########################################################################
   dest_filename = parameters[2].valueAsText;

   scenarioids = [];
   val = parameters[3].valueAsText;
   scenarioids.append(val);

   #########################################################################
   # Step 30
   # Initialize the workbook and summary sheet
   #########################################################################
   haz = source.obj_AllHazardsWasteLogisticsTool.AllHazardsWasteLogisticsTool();
   wb = Workbook();
   sum = wb.active;
   sum.title = 'Summary';

   sum['A1'] = "All Hazards Waste Logistics Tool Summary"
   ft18 = Font(size=18);
   bldu = Font(bold=True,underline="single");
   bld  = Font(bold=True);
   lft  = Alignment(horizontal='left');
   rht  = Alignment(horizontal='right');
   dol  = "$#,##0.00_);($#,##0.00)"

   sum.column_dimensions['A'].width = 20;
   sum.column_dimensions['B'].width = 30;
   sum.column_dimensions['C'].width = 30;
   sum.column_dimensions['D'].width = 22;
   sum.column_dimensions['E'].width = 27;
   sum.column_dimensions['F'].width = 18;
   sum.column_dimensions['G'].width = 28;
   sum.column_dimensions['H'].width = 18;
   sum.column_dimensions['I'].width = 33;
   sum.column_dimensions['J'].width = 14;

   sum['A1'].font = ft18;
   sum['A2'] = "Version: " + const_version;

   row_cnt = 4;

   ary_conditionids = [];
   ary_factorids     = [];

   for scenarioid in scenarioids:
      arcpy.AddMessage("  Preparing to write report for " + str(scenarioid) + ".");
      
      sc = source.obj_Scenario.Scenario(scenario_id=scenarioid);
      haz.scenario.loadScenarioID(scenarioid=scenarioid);

      if sc.conditionid not in ary_conditionids:
         ary_conditionids.append(sc.conditionid);
      if sc.factorid not in ary_factorids:
         ary_factorids.append(sc.factorid);

      total_number_of_facilities        = 0;
      total_allocated_amount            = 0;
      total_number_of_road_shipments    = 0;
      total_number_of_rail_shipments    = 0;
      total_cplm_cost_usd               = 0;
      total_fixed_cost_usd_per_contnr   = 0;
      total_fixed_cost_usd_per_hour     = 0;
      total_fixed_cost_usd_by_volume    = 0;
      total_tolls_usd                   = 0;
      total_misc_trans_cost_usd         = 0;
      total_trans_cost_usd              = 0;
      total_staging_site_cost_usd       = 0;
      total_disposal_cost_usd           = 0;
      total_labor_cost_usd              = 0;
      total_vehicle_decon_cost_usd      = 0;
      total_cost_multiplier_usd         = 0;
      total_total_cost_usd              = 0;
      max_road_transp_time_to_comp_days = 0;
      max_road_dest_time_to_comp_days   = 0;
      max_time_days                     = 0;

      unallocated_amount                = 0;
      cursor_in = arcpy.da.SearchCursor(
          in_table     = haz.scenario_results.dataSource
         ,field_names  = (
             'allocated_amount'
          )
         ,where_clause = "scenarioid = " + source.util.sql_quote(scenarioid) + " and facility_identifier = 'Unallocated'"
      );

      for row in cursor_in:
         unallocated_amount = row[0];
         
      qryset = source.obj_QuerySet.QuerySet({
          'scenarioid'                     : 'scenarioid'
         ,'conditionid'                    : 'conditionid'
         ,'factorid'                       : 'factorid'
         ,'facilityattributesid'           : 'facilityattributesid'
         ,'road_transporter_attributes'    : 'road_transporter_attributes'
         ,'rail_transporter_attributes'    : 'rail_transporter_attributes'
         ,'facility_identifier'            : 'facility_identifier'
         ,'facility_rank'                  : 'facility_rank'
         
         ,'total_overall_distance'         : 'total_overall_distance'
         ,'total_overall_distance_unt'     : 'total_overall_distance_unt'
         ,'total_overall_time'             : 'total_overall_time'
         ,'total_overall_time_unt'         : 'total_overall_time_unt'
         
         ,'total_road_distance'            : 'total_road_distance'
         ,'total_road_distance_unt'        : 'total_road_distance_unt'
         ,'total_road_time'                : 'total_road_time'
         ,'total_road_time_unt'            : 'total_road_time_unt'
         
         ,'total_rail_distance'            : 'total_rail_distance'
         ,'total_rail_distance_unt'        : 'total_rail_distance_unt'
         ,'total_rail_time'                : 'total_rail_time'
         ,'total_rail_time_unt'            : 'total_rail_time_unt'
         
         ,'total_station_count'            : 'total_station_count'
         
         ,'average_overall_speed'          : 'average_overall_speed'
         ,'average_road_speed'             : 'average_road_speed'
         ,'average_rail_speed'             : 'average_rail_speed'
         ,'speed_unit'                     : 'speed_unit'
         
         ,'facility_typeid'                : 'facility_typeid'
         ,'facility_subtypeids'            : 'facility_subtypeids'
         ,'facility_name'                  : 'facility_name'
         ,'facility_address'               : 'facility_address'
         ,'facility_city'                  : 'facility_city'
         ,'facility_state'                 : 'facility_state'
         ,'facility_zip'                   : 'facility_zip'
         ,'facility_telephone'             : 'facility_telephone'
         ,'facility_waste_mgt'             : 'facility_waste_mgt'
         
         ,'facility_dly_cap'               : 'facility_dly_cap'
         ,'facility_dly_cap_unt'           : 'facility_dly_cap_unt'
         
         ,'facility_qty_accepted'          : 'facility_qty_accepted'
         ,'facility_qty_accepted_unt'      : 'facility_qty_accepted_unt'
         
         ,'allocated_amount'               : 'allocated_amount'
         ,'allocated_amount_unt'           : 'allocated_amount_unt'
         
         ,'number_of_road_shipments'       : 'number_of_road_shipments'
         ,'number_of_rail_shipments'       : 'number_of_rail_shipments'
         
         ,'road_cplm_cost_usd'             : 'road_cplm_cost_usd'
         ,'rail_cplm_cost_usd'             : 'rail_cplm_cost_usd'
         
         ,'road_fixed_cost_usd_per_contnr' : 'road_fixed_cost_usd_per_contnr'
         ,'rail_fixed_cost_usd_per_contnr' : 'rail_fixed_cost_usd_per_contnr'
         
         ,'road_fixed_cost_usd_per_hour'   : 'road_fixed_cost_usd_per_hour'
         ,'rail_fixed_cost_usd_per_hour'   : 'rail_fixed_cost_usd_per_hour'
         
         ,'road_fixed_cost_usd_by_volume'  : 'road_fixed_cost_usd_by_volume'
         ,'rail_fixed_cost_usd_by_volume'  : 'rail_fixed_cost_usd_by_volume'
         
         ,'road_tolls_usd_per_shipment'    : 'road_tolls_usd_per_shipment'
         
         ,'road_misc_trans_cost_usd'       : 'road_misc_trans_cost_usd'
         ,'rail_misc_trans_cost_usd'       : 'rail_misc_trans_cost_usd'
         
         ,'road_trans_cost_usd'            : 'road_trans_cost_usd'
         ,'rail_trans_cost_usd'            : 'rail_trans_cost_usd'
         
         ,'staging_site_cost_usd'          : 'staging_site_cost_usd'
         
         ,'disposal_cost_usd'              : 'disposal_cost_usd'
         
         ,'road_labor_cost_usd'            : 'road_labor_cost_usd'
         ,'rail_labor_cost_usd'            : 'rail_labor_cost_usd'
         
         ,'road_transp_decon_cost_usd'     : 'road_transp_decon_cost_usd'
         ,'rail_transp_decon_cost_usd'     : 'rail_transp_decon_cost_usd'
         
         ,'cost_multiplier_usd'            : 'cost_multiplier_usd'
         
         ,'total_cost_usd'                 : 'total_cost_usd'
         
         ,'road_transp_time_to_comp_days'  : 'road_transp_time_to_comp_days'
         ,'rail_transp_time_to_comp_days'  : 'rail_transp_time_to_comp_days'
         ,'total_transp_time_to_comp_days' : 'total_transp_time_to_comp_days'
         
         ,'road_dest_time_to_comp_days'    : 'road_dest_time_to_comp_days'
         ,'rail_dest_time_to_comp_days'    : 'rail_dest_time_to_comp_days'
         ,'total_dest_time_to_comp_days'   : 'total_dest_time_to_comp_days'
         
         ,'time_days'                      : 'time_days'
         ,'username'                       : 'username'
         ,'creationtime'                   : 'creationtime'
      });

      cursor_in = arcpy.da.SearchCursor(
          in_table     = haz.scenario_results.dataSource
         ,field_names  = qryset.flds
         ,where_clause = "scenarioid = " + source.util.sql_quote(scenarioid) + " and facility_identifier <> 'Unallocated'"
      );

      sht = wb.create_sheet(scenarioid);

      sht.column_dimensions['A'].width  = 19;
      sht.column_dimensions['B'].width  = 25;
      sht.column_dimensions['C'].width  = 30;
      sht.column_dimensions['D'].width  = 25;
      sht.column_dimensions['E'].width  = 20;
      sht.column_dimensions['F'].width  = 13;
      sht.column_dimensions['G'].width  = 11;
      sht.column_dimensions['H'].width  = 18;
      sht.column_dimensions['I'].width  = 13;
      sht.column_dimensions['J'].width  = 18;
      sht.column_dimensions['K'].width  = 18;
      sht.column_dimensions['L'].width  = 10;
      #-- Total Distance
      sht.column_dimensions['M'].width  = 15;
      sht.column_dimensions['N'].width  = 15;
      sht.column_dimensions['O'].width  = 15;
      sht.column_dimensions['P'].width  = 10;
      #-- Total Travel Time
      sht.column_dimensions['Q'].width  = 18;
      sht.column_dimensions['R'].width  = 18;
      sht.column_dimensions['S'].width  = 18;
      sht.column_dimensions['T'].width  = 10;
      #-- Speed
      sht.column_dimensions['U'].width  = 20;
      sht.column_dimensions['V'].width  = 20;
      sht.column_dimensions['W'].width  = 10;     
      #-- number of road shipments
      sht.column_dimensions['X'].width  = 25;
      #-- number of rail shipments
      sht.column_dimensions['Y'].width  = 25;
      #--      
      sht.column_dimensions['Z'].width  = 18;
      sht.column_dimensions['AA'].width  = 18;
      sht.column_dimensions['AB'].width = 18;
      sht.column_dimensions['AC'].width = 20;
      sht.column_dimensions['AD'].width = 30;
      sht.column_dimensions['AE'].width = 20;
      sht.column_dimensions['AF'].width = 18;
      sht.column_dimensions['AG'].width = 19;
      sht.column_dimensions['AH'].width = 21;
      sht.column_dimensions['AI'].width = 21;
      sht.column_dimensions['AJ'].width = 27;
      sht.column_dimensions['AK'].width = 27;
      sht.column_dimensions['AL'].width = 20;
      sht.column_dimensions['AM'].width = 25;
      sht.column_dimensions['AN'].width = 25;
      sht.column_dimensions['AO'].width = 23;

      sht['A1'] = scenarioid;
      sht['A1'].font = ft18;

      sht['A3'] = "Waste Type";
      sht['A3'].font = bld;
      sht['B3'] = sc.waste_type;

      sht['A4'] = "Waste Medium";
      sht['A4'].font = bld;
      sht['B4'] = sc.waste_medium;

      sht['A5'] = "Total Waste Amount";
      sht['A5'].font = bld;
      sht['B5'] = sc.waste_amount;

      sht['A6'] = "Unallocated Amount";
      sht['A6'].font = bld;
      sht['B6'] = unallocated_amount;

      sht['A7'] = "Waste Unit";
      sht['A7'].font = bld;
      sht['B7'] = sc.waste_unit;
      sht['B7'].alignment = rht;

      sht['A9'] = "Condition ID";
      sht['A9'].font = bld;
      sht['B9'] = sc.conditionid;

      sht['A10'] = "Factor ID";
      sht['A10'].font = bld;
      sht['B10'] = sc.factorid

      if haz.scenario.map_image == 'Disabled':
         row_sht = 12;

      else:
         try:
            img = openpyxl.drawing.image.Image(
               haz.scenario.map_image
            );
            dpi = 96;
            img.width = 8 * dpi;
            img.height = 6 * dpi;
            img.anchor = 'D2';
            sht.add_image(img);

            row_sht = 32;
            
         except:
            arcpy.AddMessage(".  ERROR, cannot add image from " + str(haz.scenario.map_image));
            row_sht = 12;

      sht['A' + str(row_sht)] = "Facility ID";
      sht['A' + str(row_sht)].font = bld;

      sht['B' + str(row_sht)] = "Facility TypeIDs";
      sht['B' + str(row_sht)].font = bld;
      
      sht['C' + str(row_sht)] = "Facility Name";
      sht['C' + str(row_sht)].font = bld;
      
      sht['D' + str(row_sht)] = "Facility Address";
      sht['D' + str(row_sht)].font = bld;
      
      sht['E' + str(row_sht)] = "Facility City";
      sht['E' + str(row_sht)].font = bld;
      
      sht['F' + str(row_sht)] = "Facility State";
      sht['F' + str(row_sht)].font = bld;
      
      sht['G' + str(row_sht)] = "Facility Zip";
      sht['G' + str(row_sht)].font = bld;

      sht['H' + str(row_sht)] = "Facility Telephone";
      sht['H' + str(row_sht)].font = bld;

      sht['I' + str(row_sht)] = "Routing Rank";
      sht['I' + str(row_sht)].font = bld;

      sht['J' + str(row_sht)] = "Quantity Accepted";
      sht['J' + str(row_sht)].font = bld;

      sht['K' + str(row_sht)] = "Allocated Amount";
      sht['K' + str(row_sht)].font = bld;

      sht['L' + str(row_sht)] = "Unit";
      sht['L' + str(row_sht)].font = bld;

      sht['M' + str(row_sht)] = "Total Distance";
      sht['M' + str(row_sht)].font = bld;

      sht['N' + str(row_sht)] = "Road Distance";
      sht['N' + str(row_sht)].font = bld;
      
      sht['O' + str(row_sht)] = "Rail Distance";
      sht['O' + str(row_sht)].font = bld;
      
      sht['P' + str(row_sht)] = "Unit";
      sht['P' + str(row_sht)].font = bld;

      sht['Q' + str(row_sht)] = "Total Travel Time";
      sht['Q' + str(row_sht)].font = bld;
      
      sht['R' + str(row_sht)] = "Road Travel Time";
      sht['R' + str(row_sht)].font = bld;

      sht['S' + str(row_sht)] = "Rail Travel Time";
      sht['S' + str(row_sht)].font = bld;

      sht['T' + str(row_sht)] = "Unit";
      sht['T' + str(row_sht)].font = bld;

      sht['U' + str(row_sht)] = "Average Road Speed";
      sht['U' + str(row_sht)].font = bld;
      
      sht['V' + str(row_sht)] = "Average Rail Speed";
      sht['V' + str(row_sht)].font = bld;

      sht['W' + str(row_sht)] = "Unit";
      sht['W' + str(row_sht)].font = bld;

      sht['X' + str(row_sht)] = "Number of Road Shipments";
      sht['X' + str(row_sht)].font = bld;
      
      sht['Y' + str(row_sht)] = "Number of Rail Shipments";
      sht['Y' + str(row_sht)].font = bld;

      sht['Z' + str(row_sht)] = "CPLM Cost ($)";
      sht['Z' + str(row_sht)].font = bld;

      sht['AA' + str(row_sht)] = "Road Fixed Cost ($)";
      sht['AA' + str(row_sht)].font = bld;
      
      sht['AB' + str(row_sht)] = "Rail Fixed Cost ($)";
      sht['AB' + str(row_sht)].font = bld;

      sht['AC' + str(row_sht)] = "Tolls ($)";
      sht['AC' + str(row_sht)].font = bld;

      sht['AD' + str(row_sht)] = "Misc Trans Costs ($)";
      sht['AD' + str(row_sht)].font = bld;

      sht['AE' + str(row_sht)] = "Total Transportation Cost ($)";
      sht['AE' + str(row_sht)].font = bld;

      sht['AF' + str(row_sht)] = "Staging Site Cost ($)";
      sht['AF' + str(row_sht)].font = bld;

      sht['AG' + str(row_sht)] = "Disposal Cost ($)";
      sht['AG' + str(row_sht)].font = bld;

      sht['AH' + str(row_sht)] = "Road Labor Cost ($)";
      sht['AH' + str(row_sht)].font = bld;
      
      sht['AI' + str(row_sht)] = "Rail Labor Cost ($)";
      sht['AI' + str(row_sht)].font = bld;

      sht['AJ' + str(row_sht)] = "Vehicle Decon Cost ($)";
      sht['AJ' + str(row_sht)].font = bld;
      
      sht['AK' + str(row_sht)] = "Total Cost Multiplier ($)";
      sht['AK' + str(row_sht)].font = bld;
      
      sht['AL' + str(row_sht)] = "Total Cost ($)";
      sht['AL' + str(row_sht)].font = bld;

      sht['AM' + str(row_sht)] = "Vehicle Time to Complete (days)";
      sht['AM' + str(row_sht)].font = bld;

      sht['AN' + str(row_sht)] = "Destination Time to Complete (days)";
      sht['AN' + str(row_sht)].font = bld;

      sht['AO' + str(row_sht)] = "Total Time (days)";
      sht['AO' + str(row_sht)].font = bld;

      row_sht += 2;

      max_road_transp_time_to_comp_days = 0;
      max_road_dest_time_to_comp_days   = 0;
      max_time_days                     = 0;

      for row in cursor_in:
         conditionid                     = qryset.idx(row,'conditionid');
         factorid                        = qryset.idx(row,'factorid');
         facility_identifier             = qryset.idx(row,'facility_identifier');
         facility_rank                   = qryset.idx(row,'facility_rank');
         
         total_overall_distance          = qryset.idx(row,'total_overall_distance');
         total_overall_distance_unt      = qryset.idx(row,'total_overall_distance_unt');
         total_overall_time              = qryset.idx(row,'total_overall_time');
         total_overall_time_unt          = qryset.idx(row,'total_overall_time_unt');
         
         total_road_distance             = qryset.idx(row,'total_road_distance');
         total_road_distance_unt         = qryset.idx(row,'total_road_distance_unt');
         total_road_time                 = qryset.idx(row,'total_road_time');
         total_road_time_unt             = qryset.idx(row,'total_road_time_unt');
         
         total_rail_distance             = qryset.idx(row,'total_rail_distance');
         total_rail_distance_unt         = qryset.idx(row,'total_rail_distance_unt');
         total_rail_time                 = qryset.idx(row,'total_rail_time');
         total_rail_time_unt             = qryset.idx(row,'total_rail_time_unt');
         
         average_overall_speed           = qryset.idx(row,'average_overall_speed');
         average_road_speed              = qryset.idx(row,'average_road_speed');
         average_rail_speed              = qryset.idx(row,'average_rail_speed');
         speed_unit                      = qryset.idx(row,'speed_unit');
         
         facility_subtypeids             = qryset.idx(row,'facility_subtypeids')
         facility_name                   = qryset.idx(row,'facility_name');
         facility_address                = qryset.idx(row,'facility_address');
         facility_city                   = qryset.idx(row,'facility_city');
         facility_state                  = qryset.idx(row,'facility_state');
         facility_zip                    = qryset.idx(row,'facility_zip');
         facility_telephone              = qryset.idx(row,'facility_telephone');
         facility_waste_mgt              = qryset.idx(row,'facility_waste_mgt');
         
         facility_dly_cap                = qryset.idx(row,'facility_dly_cap');
         facility_dly_cap_unt            = qryset.idx(row,'facility_dly_cap_unt');
         
         facility_qty_accepted           = qryset.idx(row,'facility_qty_accepted');
         facility_qty_accepted_unt       = qryset.idx(row,'facility_qty_accepted_unt');
         
         allocated_amount                = qryset.idx(row,'allocated_amount');
         allocated_amount_unt            = qryset.idx(row,'allocated_amount_unt');
         
         number_of_road_shipments        = qryset.idx(row,'number_of_road_shipments');
         number_of_rail_shipments        = qryset.idx(row,'number_of_rail_shipments');
         
         road_cplm_cost_usd              = qryset.idx(row,'road_cplm_cost_usd');
         rail_cplm_cost_usd              = qryset.idx(row,'rail_cplm_cost_usd');
         
         road_fixed_cost_usd_per_contnr  = qryset.idx(row,'road_fixed_cost_usd_per_contnr');
         rail_fixed_cost_usd_per_contnr  = qryset.idx(row,'rail_fixed_cost_usd_per_contnr');
         
         road_fixed_cost_usd_per_hour    = qryset.idx(row,'road_fixed_cost_usd_per_hour');
         rail_fixed_cost_usd_per_hour    = qryset.idx(row,'rail_fixed_cost_usd_per_hour');
         
         road_fixed_cost_usd_by_volume   = qryset.idx(row,'road_fixed_cost_usd_by_volume');
         rail_fixed_cost_usd_by_volume   = qryset.idx(row,'rail_fixed_cost_usd_by_volume');
         
         road_tolls_usd_per_shipment     = qryset.idx(row,'road_tolls_usd_per_shipment');
         
         road_misc_trans_cost_usd        = qryset.idx(row,'road_misc_trans_cost_usd');
         rail_misc_trans_cost_usd        = qryset.idx(row,'rail_misc_trans_cost_usd');
         
         road_trans_cost_usd             = qryset.idx(row,'road_trans_cost_usd');
         rail_trans_cost_usd             = qryset.idx(row,'rail_trans_cost_usd');
         
         staging_site_cost_usd           = qryset.idx(row,'staging_site_cost_usd');
         
         disposal_cost_usd               = qryset.idx(row,'disposal_cost_usd');
         
         road_labor_cost_usd             = qryset.idx(row,'road_labor_cost_usd');
         rail_labor_cost_usd             = qryset.idx(row,'rail_labor_cost_usd');
         
         road_transp_decon_cost_usd      = qryset.idx(row,'road_transp_decon_cost_usd');
         rail_transp_decon_cost_usd      = qryset.idx(row,'rail_transp_decon_cost_usd');
         
         cost_multiplier_usd             = qryset.idx(row,'cost_multiplier_usd');
         
         total_cost_usd                  = qryset.idx(row,'total_cost_usd');
         
         road_transp_time_to_comp_days   = qryset.idx(row,'road_transp_time_to_comp_days');
         rail_transp_time_to_comp_days   = qryset.idx(row,'rail_transp_time_to_comp_days');
         total_transp_time_to_comp_days  = qryset.idx(row,'total_transp_time_to_comp_days');
         
         road_dest_time_to_comp_days     = qryset.idx(row,'road_dest_time_to_comp_days');
         rail_dest_time_to_comp_days     = qryset.idx(row,'rail_dest_time_to_comp_days');
         total_dest_time_to_comp_days    = qryset.idx(row,'total_dest_time_to_comp_days');
         
         time_days                       = qryset.idx(row,'time_days');
         username                        = qryset.idx(row,'username');
         creationtime                    = qryset.idx(row,'creationtime');

         total_number_of_facilities      += 1;
         total_allocated_amount          += allocated_amount;
         total_number_of_road_shipments  += number_of_road_shipments;
         total_number_of_rail_shipments  += number_of_rail_shipments;
         total_cplm_cost_usd             += road_cplm_cost_usd + rail_cplm_cost_usd;
         total_fixed_cost_usd_per_contnr += road_fixed_cost_usd_per_contnr + rail_fixed_cost_usd_per_contnr;
         total_fixed_cost_usd_per_hour   += road_fixed_cost_usd_per_hour + rail_fixed_cost_usd_per_hour;
         total_fixed_cost_usd_by_volume  += road_fixed_cost_usd_by_volume + rail_fixed_cost_usd_by_volume;
         total_tolls_usd                 += road_tolls_usd_per_shipment;
         total_misc_trans_cost_usd       += road_misc_trans_cost_usd + rail_misc_trans_cost_usd;
         total_trans_cost_usd            += road_trans_cost_usd + rail_trans_cost_usd;
         total_staging_site_cost_usd     += staging_site_cost_usd;
         total_disposal_cost_usd         += disposal_cost_usd;
         total_labor_cost_usd            += road_labor_cost_usd + rail_labor_cost_usd;
         total_vehicle_decon_cost_usd    += road_transp_decon_cost_usd + rail_transp_decon_cost_usd;
         total_cost_multiplier_usd       += cost_multiplier_usd;
         total_total_cost_usd            += total_cost_usd;

         if road_transp_time_to_comp_days > max_road_transp_time_to_comp_days:
            max_road_transp_time_to_comp_days = road_transp_time_to_comp_days;

         if road_dest_time_to_comp_days > max_road_dest_time_to_comp_days:
            max_road_dest_time_to_comp_days = road_dest_time_to_comp_days;

         if time_days is not None and time_days > max_time_days:
            max_time_days = time_days;

         sht['A' + str(row_sht)] = facility_identifier;

         sht['B' + str(row_sht)] = waste.subtypeids2txt(facility_subtypeids);
         
         sht['C' + str(row_sht)] = facility_name;
         
         sht['D' + str(row_sht)] = facility_address;
         
         sht['E' + str(row_sht)] = facility_city;
         
         sht['F' + str(row_sht)] = facility_state;
         
         sht['G' + str(row_sht)] = facility_zip;
         
         sht['H' + str(row_sht)] = facility_telephone;
 
         sht['I' + str(row_sht)] = facility_rank;

         sht['J' + str(row_sht)] = facility_qty_accepted;

         sht['K' + str(row_sht)] = allocated_amount;

         sht['L' + str(row_sht)] = allocated_amount_unt;

         sht['M' + str(row_sht)] = total_overall_distance;
         sht['N' + str(row_sht)] = total_road_distance;
         sht['O' + str(row_sht)] = total_rail_distance;
         sht['P' + str(row_sht)] = total_overall_distance_unt;

         sht['Q' + str(row_sht)] = total_overall_time;
         sht['R' + str(row_sht)] = total_road_time;
         sht['S' + str(row_sht)] = total_rail_time;
         sht['T' + str(row_sht)] = total_overall_time_unt;

         sht['U' + str(row_sht)] = average_road_speed;
         sht['V' + str(row_sht)] = average_rail_speed;
         sht['W' + str(row_sht)] = speed_unit;

         sht['X' + str(row_sht)] = number_of_road_shipments;
         
         sht['Y' + str(row_sht)] = number_of_rail_shipments;

         sht['Z' + str(row_sht)] = road_cplm_cost_usd;
         sht['Z' + str(row_sht)].number_format = dol;

         sht['AA' + str(row_sht)] = road_fixed_cost_usd_per_contnr + road_fixed_cost_usd_per_hour + road_fixed_cost_usd_by_volume;
         sht['AA' + str(row_sht)].number_format = dol;
         
         sht['AB' + str(row_sht)] = rail_fixed_cost_usd_per_contnr + rail_fixed_cost_usd_per_hour + rail_fixed_cost_usd_by_volume;
         sht['AB' + str(row_sht)].number_format = dol;

         sht['AC' + str(row_sht)] = road_tolls_usd_per_shipment;
         sht['AC' + str(row_sht)].number_format = dol;

         sht['AD' + str(row_sht)] = road_misc_trans_cost_usd;
         sht['AD' + str(row_sht)].number_format = dol;

         sht['AE' + str(row_sht)] = road_trans_cost_usd;
         sht['AE' + str(row_sht)].number_format = dol;

         sht['AF' + str(row_sht)] = staging_site_cost_usd;
         sht['AF' + str(row_sht)].number_format = dol;

         sht['AG' + str(row_sht)] = disposal_cost_usd;
         sht['AG' + str(row_sht)].number_format = dol;

         sht['AH' + str(row_sht)] = road_labor_cost_usd;
         sht['AH' + str(row_sht)].number_format = dol;
         
         sht['AI' + str(row_sht)] = rail_labor_cost_usd;
         sht['AI' + str(row_sht)].number_format = dol;

         sht['AJ' + str(row_sht)] = road_transp_decon_cost_usd;
         sht['AJ' + str(row_sht)].number_format = dol;

         sht['AK' + str(row_sht)] = cost_multiplier_usd;
         sht['AK' + str(row_sht)].number_format = dol;
         
         sht['AL' + str(row_sht)] = total_cost_usd;
         sht['AL' + str(row_sht)].number_format = dol;

         sht['AM' + str(row_sht)] = road_transp_time_to_comp_days;

         sht['AN' + str(row_sht)] = road_dest_time_to_comp_days;

         sht['AO' + str(row_sht)] = time_days;

         row_sht += 1;

      sum['A' + str(row_cnt)] = "ScenarioID";
      sum['A' + str(row_cnt)].font = bldu;
      sum['B' + str(row_cnt)] = scenarioid;

      sum['A' + str(row_cnt + 1)] = "Waste Type";
      sum['A' + str(row_cnt + 1)].font = bld;
      sum['B' + str(row_cnt + 1)] = sc.waste_type;

      sum['A' + str(row_cnt + 2)] = "Waste Medium";
      sum['A' + str(row_cnt + 2)].font = bld;
      sum['B' + str(row_cnt + 2)] = sc.waste_medium;

      sum['A' + str(row_cnt + 3)] = "Total Waste Amount";
      sum['A' + str(row_cnt + 3)].font = bld;
      sum['B' + str(row_cnt + 3)] = sc.waste_amount;

      sum['A' + str(row_cnt + 4)] = "Allocated Amount";
      sum['A' + str(row_cnt + 4)].font = bld;
      sum['B' + str(row_cnt + 4)] = total_allocated_amount;

      sum['A' + str(row_cnt + 5)] = "Unallocated Amount";
      sum['A' + str(row_cnt + 5)].font = bld;
      sum['B' + str(row_cnt + 5)] = unallocated_amount;

      sum['A' + str(row_cnt + 6)] = "Waste Unit";
      sum['A' + str(row_cnt + 6)].font = bld;
      sum['B' + str(row_cnt + 6)] = sc.waste_unit;
      sum['B' + str(row_cnt + 6)].alignment = rht;

      sum['C' + str(row_cnt + 1)] = "ConditionID";
      sum['C' + str(row_cnt + 1)].font = bld;
      sum['D' + str(row_cnt + 1)] = sc.conditionid;

      sum['C' + str(row_cnt + 2)] = "FactorID";
      sum['C' + str(row_cnt + 2)].font = bld;
      sum['D' + str(row_cnt + 2)] = sc.factorid;
      
      sum['C' + str(row_cnt + 3)] = "Facility Attributes ID";
      sum['C' + str(row_cnt + 3)].font = bld;
      sum['D' + str(row_cnt + 3)] = sc.facilityattributesid;
      
      sum['C' + str(row_cnt + 4)] = "Road Transporter Attrs ID";
      sum['C' + str(row_cnt + 4)].font = bld;
      sum['D' + str(row_cnt + 4)] = sc.road_transporter_attributes;
      
      sum['C' + str(row_cnt + 5)] = "Rail Transporter Attrs ID";
      sum['C' + str(row_cnt + 5)].font = bld;
      sum['D' + str(row_cnt + 5)] = sc.rail_transporter_attributes;

      sum['C' + str(row_cnt + 6)] = "Total Number of Facilities";
      sum['C' + str(row_cnt + 6)].font = bld;
      sum['D' + str(row_cnt + 6)] = total_number_of_facilities;

      sum['C' + str(row_cnt + 7)] = "Total Number of Road Shipments";
      sum['C' + str(row_cnt + 7)].font = bld;
      sum['D' + str(row_cnt + 7)] = total_number_of_road_shipments;
      
      sum['C' + str(row_cnt + 8)] = "Total Number of Rail Shipments";
      sum['C' + str(row_cnt + 8)].font = bld;
      sum['D' + str(row_cnt + 8)] = total_number_of_rail_shipments;

      sum['E' + str(row_cnt + 1)] = "Total CPLM Cost ($)";
      sum['E' + str(row_cnt + 1)].font = bld;
      sum['F' + str(row_cnt + 1)] = total_cplm_cost_usd;
      sum['F' + str(row_cnt + 1)].number_format = dol;

      sum['E' + str(row_cnt + 2)] = "Total Fixed Cost Per Contnr ($)";
      sum['E' + str(row_cnt + 2)].font = bld;
      sum['F' + str(row_cnt + 2)] = total_fixed_cost_usd_per_contnr;
      sum['F' + str(row_cnt + 2)].number_format = dol;
      
      sum['E' + str(row_cnt + 3)] = "Total Fixed Cost Per Hour ($)";
      sum['E' + str(row_cnt + 3)].font = bld;
      sum['F' + str(row_cnt + 3)] = total_fixed_cost_usd_per_hour;
      sum['F' + str(row_cnt + 3)].number_format = dol;
      
      sum['E' + str(row_cnt + 4)] = "Total Fixed Cost By Volume ($)";
      sum['E' + str(row_cnt + 4)].font = bld;
      sum['F' + str(row_cnt + 4)] = total_fixed_cost_usd_by_volume;
      sum['F' + str(row_cnt + 4)].number_format = dol;

      sum['E' + str(row_cnt + 5)] = "Total Tolls ($)";
      sum['E' + str(row_cnt + 5)].font = bld;
      sum['F' + str(row_cnt + 5)] = total_tolls_usd;
      sum['F' + str(row_cnt + 5)].number_format = dol;

      sum['E' + str(row_cnt + 6)] = "Total Misc Trans Costs ($)";
      sum['E' + str(row_cnt + 6)].font = bld;
      sum['F' + str(row_cnt + 6)] = total_misc_trans_cost_usd;
      sum['F' + str(row_cnt + 6)].number_format = dol;

      sum['E' + str(row_cnt + 7)] = "Total Transportation Cost ($)";
      sum['E' + str(row_cnt + 7)].font = bld;
      sum['F' + str(row_cnt + 7)] = total_trans_cost_usd;
      sum['F' + str(row_cnt + 7)].number_format = dol;

      sum['G' + str(row_cnt + 1)] = "Total Staging Site Cost ($)";
      sum['G' + str(row_cnt + 1)].font = bld;
      sum['H' + str(row_cnt + 1)] = total_staging_site_cost_usd;
      sum['H' + str(row_cnt + 1)].number_format = dol;

      sum['G' + str(row_cnt + 2)] = "Total Disposal Cost ($)";
      sum['G' + str(row_cnt + 2)].font = bld;
      sum['H' + str(row_cnt + 2)] = total_disposal_cost_usd;
      sum['H' + str(row_cnt + 2)].number_format = dol;

      sum['G' + str(row_cnt + 3)] = "Total Labor Cost ($)";
      sum['G' + str(row_cnt + 3)].font = bld;
      sum['H' + str(row_cnt + 3)] = total_labor_cost_usd;
      sum['H' + str(row_cnt + 3)].number_format = dol;

      sum['G' + str(row_cnt + 4)] = "Total Vehicle Decon Cost ($)";
      sum['G' + str(row_cnt + 4)].font = bld;
      sum['H' + str(row_cnt + 4)] = total_vehicle_decon_cost_usd;
      sum['H' + str(row_cnt + 4)].number_format = dol;

      sum['G' + str(row_cnt + 5)] = "Total Cost Multiplier ($)";
      sum['G' + str(row_cnt + 5)].font = bld;
      sum['H' + str(row_cnt + 5)] = total_cost_multiplier_usd;
      sum['H' + str(row_cnt + 5)].number_format = dol;
      
      #
      
      sum['G' + str(row_cnt + 7)] = "Total Cost ($)";
      sum['G' + str(row_cnt + 7)].font = bld;
      sum['H' + str(row_cnt + 7)] = total_total_cost_usd;
      sum['H' + str(row_cnt + 7)].number_format = dol;

      sum['I' + str(row_cnt + 1)] = "Vehicle Time to Complete (days)";
      sum['I' + str(row_cnt + 1)].font = bld;
      sum['J' + str(row_cnt + 1)] = max_road_transp_time_to_comp_days;

      sum['I' + str(row_cnt + 2)] = "Destination Time to Complete (days)";
      sum['I' + str(row_cnt + 2)].font = bld;
      sum['J' + str(row_cnt + 2)] = max_road_dest_time_to_comp_days;

      sum['I' + str(row_cnt + 3)] = "Total Time Days (days)";
      sum['I' + str(row_cnt + 3)].font = bld;
      sum['J' + str(row_cnt + 3)] = max_time_days;

      row_cnt =+ 8;

   #########################################################################
   # Step 40
   # Add the Factor sheet
   #########################################################################
   ref = wb.create_sheet('Reference');

   ref.column_dimensions['A'].width = 12;
   ref.column_dimensions['B'].width = 38;
   ref.column_dimensions['C'].width = 25;
   ref.column_dimensions['D'].width = 25;
   ref.column_dimensions['E'].width = 30;
   ref.column_dimensions['F'].width = 25;
   ref.column_dimensions['G'].width = 25;
   ref.column_dimensions['H'].width = 25;
   ref.column_dimensions['I'].width = 25;

   ref['A1'] = 'Reference';
   ref['A1'].font = ft18;

   row_cnt = 3;

   for cid in ary_conditionids:
      if cid is not None:
         haz.conditions.loadConditionID(cid);

         ref['A' + str(row_cnt)] = "Condition ID";
         ref['A' + str(row_cnt)].font = bld;
         ref['B' + str(row_cnt)] = cid;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Road Tolls ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.roadtollsperroadshipment;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Misc Road Costs ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.misccostperroadshipment;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = 'Misc Rail Costs ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.misccostperrailshipment;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Road Transporter Decon Cost ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.roadtransporterdeconcost;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = 'Rail Transporter Decon Cost ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.railtransporterdeconcost;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Staging Site Cost ($/day)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.stagingsitecost;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = 'Road Driving Hours (hrs/day)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.roaddrivinghoursperday;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = 'Rail Driving Hours (hrs/day)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.raildrivinghoursperday;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = 'Total Cost Multiplier (addt\'l% of total cost)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = str(haz.conditions.totalcostmultiplier * 100) + "%";
         ref['C' + str(row_cnt)].alignment = rht;
         row_cnt += 1;   

   row_cnt += 1;

   for fid in ary_factorids:
      if fid is not None:
         haz.factors.loadFactorID(fid);

         ref['A' + str(row_cnt)] = "Factor ID";
         ref['A' + str(row_cnt)].font = bld;
         ref['B' + str(row_cnt)] = fid;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = "Modes";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;
         
         ref['B' + str(row_cnt)] = "Name";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Description";
         ref['C' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.modes:
            ref['B' + str(row_cnt)] = item.name;
            ref['C' + str(row_cnt)] = item.description;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Transporters";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Transporter Attrs ID";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Mode";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Waste Type";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Waste Medium";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "Container Capacity";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "Unit";
         ref['G' + str(row_cnt)].font = bld;
         ref['H' + str(row_cnt)] = "Container Count";
         ref['H' + str(row_cnt)].font = bld;
         ref['I' + str(row_cnt)] = "Transporters Available";
         ref['I' + str(row_cnt)].font = bld;
         ref['J' + str(row_cnt)] = "transporters Processed Per Day";
         ref['J' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.transporters:
            ref['B' + str(row_cnt)] = item.transporterattrid;
            ref['C' + str(row_cnt)] = item.mode;
            ref['D' + str(row_cnt)] = item.wastetype;
            ref['E' + str(row_cnt)] = item.wastemedium;
            ref['F' + str(row_cnt)] = item.containercapacity;
            ref['G' + str(row_cnt)] = item.containercapacityunit;
            ref['H' + str(row_cnt)] = item.containercountpertransporter;
            ref['I' + str(row_cnt)] = item.transportersavailable;
            ref['J' + str(row_cnt)] = item.transportersprocessedperday;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "CPLM Unit Rates";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Mode";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "CPLMDist Lower";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "CPLMDist Upper";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Unit";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "Waste Type";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "Waste Medium";
         ref['G' + str(row_cnt)].font = bld;
         ref['H' + str(row_cnt)] = "CPLMUnit Rate";
         ref['H' + str(row_cnt)].font = bld;
         ref['I' + str(row_cnt)] = "Unit";
         ref['I' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.cplm_unit_rates:
            ref['B' + str(row_cnt)] = item.mode;
            ref['C' + str(row_cnt)] = item.cplmdist_lower;
            ref['D' + str(row_cnt)] = item.cplmdist_upper;
            ref['E' + str(row_cnt)] = item.cplmunit
            ref['F' + str(row_cnt)] = item.wastetype;
            ref['G' + str(row_cnt)] = item.wastemedium;
            ref['H' + str(row_cnt)] = item.cplunit_rate;
            ref['I' + str(row_cnt)] = item.cplunit_rateunit;
            ref['I' + str(row_cnt)].alignment = rht;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Fixed Trans Cost";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Mode";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "FixedCost Type";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Waste Type";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Waste Medium";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "FixedCost Value";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "Unit";
         ref['G' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.fixed_trans_cost:
            ref['B' + str(row_cnt)] = item.mode;
            ref['C' + str(row_cnt)] = item.fixedcost_type;
            ref['D' + str(row_cnt)] = item.wastetype;
            ref['E' + str(row_cnt)] = item.wastemedium;
            ref['F' + str(row_cnt)] = item.fixedcost_value;
            ref['G' + str(row_cnt)] = item.fixedcost_valueunit;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Labor Costs";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Mode";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Labor Category";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Labor Cost";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Unit";
         ref['E' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.labor_costs:
            ref['B' + str(row_cnt)] = item.mode;
            ref['C' + str(row_cnt)] = item.laborcategory;
            ref['D' + str(row_cnt)] = item.laborcost;
            ref['E' + str(row_cnt)] = item.laborcostunit;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Facility Capacity";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Facility Attributes ID";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Facility TypeID";
         ref['C' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)].alignment = lft;
         ref['D' + str(row_cnt)] = "Facility Type Name";
         ref['D' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)].alignment = lft;  
         ref['E' + str(row_cnt)] = "Waste Type";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "Waste Medium";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "Daily Volume";
         ref['G' + str(row_cnt)].font = bld;
         ref['H' + str(row_cnt)] = "Unit";
         ref['H' + str(row_cnt)].font = bld;
         ref['I' + str(row_cnt)] = "Total Accepted Days";
         ref['I' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.facility_capacities:
            ref['B' + str(row_cnt)] = item.facilityattributesid;
            ref['C' + str(row_cnt)] = item.facility_subtypeid;
            ref['C' + str(row_cnt)].alignment = lft;
            ref['D' + str(row_cnt)] = waste.subtypeids2txt(item.facility_subtypeid);
            ref['E' + str(row_cnt)] = item.wastetype;
            ref['F' + str(row_cnt)] = item.wastemedium;
            ref['G' + str(row_cnt)] = item.dailyvolumeperday;
            ref['H' + str(row_cnt)] = item.dailyvolumeperdayunit;
            ref['I' + str(row_cnt)] = item.totalaccepted_days;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Disposal Fees";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Facility Attributes ID";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Facility TypeID";
         ref['C' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)].alignment = lft;
         ref['D' + str(row_cnt)] = "Facility Type Name";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Waste Type";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "Waste Medium";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "Disposal Cost";
         ref['G' + str(row_cnt)].font = bld;
         ref['H' + str(row_cnt)] = "Unit";
         ref['H' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.disposal_fees:
            ref['B' + str(row_cnt)] = item.facilityattributesid;
            ref['C' + str(row_cnt)] = item.facility_subtypeid;
            ref['C' + str(row_cnt)].alignment = lft;
            ref['D' + str(row_cnt)] = waste.subtypeids2txt(item.facility_subtypeid);
            ref['E' + str(row_cnt)] = item.wastetype;
            ref['F' + str(row_cnt)] = item.wastemedium;
            ref['G' + str(row_cnt)] = item.costperone;
            ref['H' + str(row_cnt)] = item.costperoneunit;
            row_cnt += 1;

         row_cnt += 1;

   #########################################################################
   # Step 50
   # Write out the excel file
   #########################################################################
   wb.save(dest_filename);

   #########################################################################
   # Step 60
   #
   #########################################################################
   del wb,sum,sht;
   del cursor_in;
   del haz,sc;

   return;
