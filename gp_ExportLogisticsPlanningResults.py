import arcpy,os,sys;
import json;
import openpyxl;
from openpyxl import Workbook;
from openpyxl.styles import Font,Alignment;

import util;
import obj_AllHazardsWasteLogisticsTool;
import obj_Scenario;

###############################################################################
import importlib
importlib.reload(util);
importlib.reload(obj_AllHazardsWasteLogisticsTool);
importlib.reload(obj_Scenario);

const_version = "1.0";

def execute(self, parameters, messages):

   #########################################################################
   # Step 10
   # Abend if edits are pending
   #########################################################################
   if util.sniff_editing_state():
      raise arcpy.ExecuteError("Error.  Pending edits must be saved or cleared before proceeding.");
      
   #########################################################################
   # Step 20
   # Read the parameters
   #########################################################################
   dest_filename = parameters[2].valueAsText;

   scenarioids = [];
   val = parameters[3].valueAsText;
   scenarioids.append(val);

   #########################################################################
   # Step 30
   # Initialize the workbook and summary sheet
   #########################################################################
   haz = obj_AllHazardsWasteLogisticsTool.AllHazardsWasteLogisticsTool();
   wb = Workbook();
   sum = wb.active;
   sum.title = 'Summary';

   sum['A1'] = "All Hazards Waste Logistics Tool Summary"
   ft18 = Font(size=18);
   bldu = Font(bold=True,underline="single");
   bld  = Font(bold=True);
   lft  = Alignment(horizontal='left');
   rht  = Alignment(horizontal='right');
   dol  = "$#,##0.00_);($#,##0.00)"

   sum.column_dimensions['A'].width = 20;
   sum.column_dimensions['B'].width = 30;
   sum.column_dimensions['C'].width = 25;
   sum.column_dimensions['D'].width = 20;
   sum.column_dimensions['E'].width = 30;
   sum.column_dimensions['F'].width = 20;
   sum.column_dimensions['G'].width = 30;
   sum.column_dimensions['H'].width = 18;
   sum.column_dimensions['I'].width = 30;
   sum.column_dimensions['J'].width = 15;

   sum['A1'].font = ft18;
   sum['A2'] = "Version: " + const_version;

   row_cnt = 4;

   ary_conditionids = [];
   ary_factorids     = [];

   for scenarioid in scenarioids:
      arcpy.AddMessage("Preparing to write report for " + str(scenarioid) + ".");
      
      sc = obj_Scenario.Scenario(scenario_id=scenarioid);

      if sc.conditionid not in ary_conditionids:
         ary_conditionids.append(sc.conditionid);
      if sc.factorid not in ary_factorids:
         ary_factorids.append(sc.factorid);

      total_number_of_facilities     = 0;
      total_allocated_amount         = 0;
      total_number_of_shipments      = 0;
      total_cplm_cost_usd            = 0;
      total_fixed_cost_usd           = 0;
      total_tolls_usd                = 0;
      total_misc_trans_cost_usd      = 0;
      total_trans_cost_usd           = 0;
      total_staging_site_cost_usd    = 0;
      total_disposal_cost_usd        = 0;
      total_labor_cost_usd           = 0;
      total_vehicle_decon_cost_usd   = 0;
      total_cost_multiplier_usd      = 0;
      total_cost_usd                 = 0;
      max_trucks_time_to_comp_days   = 0;
      max_dest_time_to_comp_days     = 0;
      max_time_days                  = 0;

      unallocated_amount             = 0;
      cursor_in = arcpy.da.SearchCursor(
          in_table     = haz.scenario_results.dataSource
         ,field_names  = (
             'allocated_amount'
          )
         ,where_clause = "scenarioid = " + util.sql_quote(scenarioid) + " and facility_identifier = 'Unallocated'"
      );

      for row in cursor_in:
         unallocated_amount = row[0];

      cursor_in = arcpy.da.SearchCursor(
          in_table     = haz.scenario_results.dataSource
         ,field_names  = (
             'scenarioid'
            ,'conditionid'
            ,'factorid'
            ,'facility_identifier'
            ,'facility_rank'
            ,'total_distance'
            ,'distance_unit'
            ,'total_truck_travel_time'
            ,'time_unit'
            ,'average_speed'
            ,'speed_unit'
            ,'facility_name'
            ,'facility_address'
            ,'facility_city'
            ,'facility_state'
            ,'facility_zip'
            ,'facility_telephone'
            ,'facility_waste_mgt'
            ,'facility_capacity_trucks_perday'
            ,'facility_qty_accepted'
            ,'facility_qty_accepted_unit'
            ,'allocated_amount'
            ,'allocated_amount_unit'
            ,'number_of_shipments'
            ,'cplm_cost_usd'
            ,'fixed_cost_usd_per_shipment'
            ,'fixed_cost_usd_per_hour'
            ,'tolls_usd'
            ,'misc_trans_cost_usd'
            ,'trans_cost_usd'
            ,'staging_site_cost_usd'
            ,'disposal_cost_usd'
            ,'labor_cost_usd'
            ,'vehicle_decon_cost_usd'
            ,'cost_multiplier_usd'
            ,'cost_usd'
            ,'trucks_time_to_comp_days'
            ,'dest_time_to_comp_days'
            ,'time_days'
            ,'username'
            ,'creationtime'
          )
         ,where_clause = "scenarioid = " + util.sql_quote(scenarioid) + " and facility_identifier <> 'Unallocated'"
      );

      sht = wb.create_sheet(scenarioid);

      sht.column_dimensions['A'].width  = 19;
      sht.column_dimensions['B'].width  = 40;
      sht.column_dimensions['C'].width  = 25;
      sht.column_dimensions['D'].width  = 20;
      sht.column_dimensions['E'].width  = 10;
      sht.column_dimensions['F'].width  = 10;
      sht.column_dimensions['G'].width  = 20;
      sht.column_dimensions['H'].width  = 15;
      sht.column_dimensions['I'].width  = 20;
      sht.column_dimensions['J'].width  = 20;
      sht.column_dimensions['K'].width  = 12;
      sht.column_dimensions['L'].width  = 15;
      sht.column_dimensions['M'].width  = 10;
      sht.column_dimensions['N'].width  = 16;
      sht.column_dimensions['O'].width  = 10;
      sht.column_dimensions['P'].width  = 15;
      sht.column_dimensions['Q'].width  = 10;
      sht.column_dimensions['R'].width  = 25;
      sht.column_dimensions['S'].width  = 18;
      sht.column_dimensions['T'].width  = 18;
      sht.column_dimensions['U'].width  = 18;
      sht.column_dimensions['V'].width  = 20;
      sht.column_dimensions['W'].width  = 30;
      sht.column_dimensions['X'].width  = 20;
      sht.column_dimensions['Y'].width  = 18;
      sht.column_dimensions['Z'].width  = 18;
      sht.column_dimensions['AA'].width = 25;
      sht.column_dimensions['AB'].width = 25;
      sht.column_dimensions['AC'].width = 18;
      sht.column_dimensions['AD'].width = 25;
      sht.column_dimensions['AE'].width = 25;
      sht.column_dimensions['AF'].width = 20;

      sht['A1'] = scenarioid;
      sht['A1'].font = ft18;

      sht['A3'] = "Waste Type";
      sht['A3'].font = bld;
      sht['B3'] = sc.waste_type;

      sht['A4'] = "Waste Medium";
      sht['A4'].font = bld;
      sht['B4'] = sc.waste_medium;

      sht['A5'] = "Total Waste Amount";
      sht['A5'].font = bld;
      sht['B5'] = sc.waste_amount;

      sht['A6'] = "Unallocated Amount";
      sht['A6'].font = bld;
      sht['B6'] = unallocated_amount;

      sht['A7'] = "Waste Unit";
      sht['A7'].font = bld;
      sht['B7'] = sc.waste_unit;
      sht['B7'].alignment = rht;

      sht['A9'] = "Condition ID";
      sht['A9'].font = bld;
      sht['B9'] = sc.conditionid;

      sht['A10'] = "Factor ID";
      sht['A10'].font = bld;
      sht['B10'] = sc.factorid

      if haz.scenario.map_image == 'Disabled':
         row_sht = 12;

      else:
         img = openpyxl.drawing.image.Image(
            haz.scenario.map_image
         );
         dpi = 96;
         img.width = 8 * dpi;
         img.height = 6 * dpi;
         img.anchor = 'D2';
         sht.add_image(img);

         row_sht = 32;

      sht['A' + str(row_sht)] = "Facility ID";
      sht['A' + str(row_sht)].font = bld;

      sht['B' + str(row_sht)] = "Facility Name";
      sht['B' + str(row_sht)].font = bld;
      
      sht['C' + str(row_sht)] = "Facility Address";
      sht['C' + str(row_sht)].font = bld;
      
      sht['D' + str(row_sht)] = "Facility City";
      sht['D' + str(row_sht)].font = bld;
      
      sht['E' + str(row_sht)] = "Facility State";
      sht['E' + str(row_sht)].font = bld;
      
      sht['F' + str(row_sht)] = "Facility Zip";
      sht['F' + str(row_sht)].font = bld;

      sht['G' + str(row_sht)] = "Facility Telephone";
      sht['G' + str(row_sht)].font = bld;

      sht['H' + str(row_sht)] = "Routing Rank";
      sht['H' + str(row_sht)].font = bld;

      sht['I' + str(row_sht)] = "Quantity Accepted";
      sht['I' + str(row_sht)].font = bld;

      sht['J' + str(row_sht)] = "Allocated Amount";
      sht['J' + str(row_sht)].font = bld;

      sht['K' + str(row_sht)] = "Unit";
      sht['K' + str(row_sht)].font = bld;

      sht['L' + str(row_sht)] = "Distance";
      sht['L' + str(row_sht)].font = bld;

      sht['M' + str(row_sht)] = "Unit";
      sht['M' + str(row_sht)].font = bld;

      sht['N' + str(row_sht)] = "Travel Time";
      sht['N' + str(row_sht)].font = bld;

      sht['O' + str(row_sht)] = "Unit";
      sht['O' + str(row_sht)].font = bld;

      sht['P' + str(row_sht)] = "Average Speed";
      sht['P' + str(row_sht)].font = bld;

      sht['Q' + str(row_sht)] = "Unit";
      sht['Q' + str(row_sht)].font = bld;

      sht['R' + str(row_sht)] = "Number of Shipments";
      sht['R' + str(row_sht)].font = bld;

      sht['S' + str(row_sht)] = "CPLM Cost ($)";
      sht['S' + str(row_sht)].font = bld;

      sht['T' + str(row_sht)] = "Fixed Cost ($)";
      sht['T' + str(row_sht)].font = bld;

      sht['U' + str(row_sht)] = "Tolls ($)";
      sht['U' + str(row_sht)].font = bld;

      sht['V' + str(row_sht)] = "Misc Trans Costs ($)";
      sht['V' + str(row_sht)].font = bld;

      sht['W' + str(row_sht)] = "Total Transportation Cost ($)";
      sht['W' + str(row_sht)].font = bld;

      sht['X' + str(row_sht)] = "Staging Site Cost ($)";
      sht['X' + str(row_sht)].font = bld;

      sht['Y' + str(row_sht)] = "Disposal Cost ($)";
      sht['Y' + str(row_sht)].font = bld;

      sht['Z' + str(row_sht)] = "Labor Cost ($)";
      sht['Z' + str(row_sht)].font = bld;

      sht['AA' + str(row_sht)] = "Vehicle Decon Cost ($)";
      sht['AA' + str(row_sht)].font = bld;
      
      sht['AB' + str(row_sht)] = "Total Cost Multiplier ($)";
      sht['AB' + str(row_sht)].font = bld;

      sht['AC' + str(row_sht)] = "Total Cost ($)";
      sht['AC' + str(row_sht)].font = bld;

      sht['AD' + str(row_sht)] = "Truck Time to Complete (days)";
      sht['AD' + str(row_sht)].font = bld;

      sht['AE' + str(row_sht)] = "Destination Time to Complete (days)";
      sht['AE' + str(row_sht)].font = bld;

      sht['AF' + str(row_sht)] = "Total Time (days)";
      sht['AF' + str(row_sht)].font = bld;

      row_sht += 2;

      max_trucks_time_to_comp_days = 0;
      max_dest_time_to_comp_days   = 0;
      max_time_days                = 0;

      for row in cursor_in:
         conditionid                     = row[1];
         factorid                        = row[2];
         facility_identifier             = row[3];
         facility_rank                   = row[4];
         total_distance                  = row[5];
         distance_unit                   = row[6];
         total_truck_travel_time         = row[7];
         time_unit                       = row[8];
         average_speed                   = row[9];
         speed_unit                      = row[10];
         facility_name                   = row[11];
         facility_address                = row[12];
         facility_city                   = row[13];
         facility_state                  = row[14];
         facility_zip                    = row[15];
         facility_telephone              = row[16];
         facility_waste_mgt              = row[17];
         facility_capacity_trucks_perday = row[18];
         facility_qty_accepted           = row[19];
         facility_qty_accepted_unit      = row[20];
         allocated_amount                = row[21];
         allocated_amount_unit           = row[22];
         number_of_shipments             = row[23];
         cplm_cost_usd                   = row[24];
         fixed_cost_usd_per_shipment     = row[25];
         fixed_cost_usd_per_hour         = row[26];
         tolls_usd                       = row[27];
         misc_trans_cost_usd             = row[28];
         trans_cost_usd                  = row[29];
         staging_site_cost_usd           = row[30];
         disposal_cost_usd               = row[31];
         labor_cost_usd                  = row[32];
         vehicle_decon_cost_usd          = row[33];
         cost_multiplier_usd             = row[34];
         cost_usd                        = row[35];
         trucks_time_to_comp_days        = row[36];
         dest_time_to_comp_days          = row[37];
         time_days                       = row[38];
         username                        = row[39];
         creationtime                    = row[40];

         total_number_of_facilities     += 1;
         total_allocated_amount         += allocated_amount;
         total_number_of_shipments      += number_of_shipments;
         total_cplm_cost_usd            += cplm_cost_usd;
         total_fixed_cost_usd           += fixed_cost_usd_per_shipment;
         total_fixed_cost_usd           += fixed_cost_usd_per_hour;
         total_tolls_usd                += tolls_usd;
         total_misc_trans_cost_usd      += misc_trans_cost_usd;
         total_trans_cost_usd           += trans_cost_usd;
         total_staging_site_cost_usd    += staging_site_cost_usd;
         total_disposal_cost_usd        += disposal_cost_usd;
         total_labor_cost_usd           += labor_cost_usd;
         total_vehicle_decon_cost_usd   += vehicle_decon_cost_usd;
         total_cost_multiplier_usd      += cost_multiplier_usd;
         total_cost_usd                 += cost_usd;

         if trucks_time_to_comp_days > max_trucks_time_to_comp_days:
            max_trucks_time_to_comp_days = trucks_time_to_comp_days;

         if dest_time_to_comp_days > max_dest_time_to_comp_days:
            max_dest_time_to_comp_days = dest_time_to_comp_days;

         if time_days > max_time_days:
            max_time_days = time_days;

         sht['A' + str(row_sht)] = facility_identifier;

         sht['B' + str(row_sht)] = facility_name;
         
         sht['C' + str(row_sht)] = facility_address;
         
         sht['D' + str(row_sht)] = facility_city;
         
         sht['E' + str(row_sht)] = facility_state;
         
         sht['F' + str(row_sht)] = facility_zip;
         
         sht['G' + str(row_sht)] = facility_telephone;
 
         sht['H' + str(row_sht)] = facility_rank;

         sht['I' + str(row_sht)] = facility_qty_accepted;

         sht['J' + str(row_sht)] = allocated_amount;

         sht['K' + str(row_sht)] = allocated_amount_unit;

         sht['L' + str(row_sht)] = total_distance;

         sht['M' + str(row_sht)] = distance_unit;

         sht['N' + str(row_sht)] = total_truck_travel_time;

         sht['O' + str(row_sht)] = time_unit;

         sht['P' + str(row_sht)] = average_speed;

         sht['Q' + str(row_sht)] = speed_unit;

         sht['R' + str(row_sht)] = number_of_shipments;

         sht['S' + str(row_sht)] = cplm_cost_usd;
         sht['S' + str(row_sht)].number_format = dol;

         sht['T' + str(row_sht)] = fixed_cost_usd_per_shipment + fixed_cost_usd_per_hour;
         sht['T' + str(row_sht)].number_format = dol;

         sht['U' + str(row_sht)] = tolls_usd;
         sht['U' + str(row_sht)].number_format = dol;

         sht['V' + str(row_sht)] = misc_trans_cost_usd;
         sht['V' + str(row_sht)].number_format = dol;

         sht['W' + str(row_sht)] = trans_cost_usd;
         sht['W' + str(row_sht)].number_format = dol;

         sht['X' + str(row_sht)] = staging_site_cost_usd;
         sht['X' + str(row_sht)].number_format = dol;

         sht['Y' + str(row_sht)] = disposal_cost_usd;
         sht['Y' + str(row_sht)].number_format = dol;

         sht['Z' + str(row_sht)] = labor_cost_usd;
         sht['Z' + str(row_sht)].number_format = dol;

         sht['AA' + str(row_sht)] = vehicle_decon_cost_usd;
         sht['AA' + str(row_sht)].number_format = dol;

         sht['AB' + str(row_sht)] = cost_multiplier_usd;
         sht['AB' + str(row_sht)].number_format = dol;
         
         sht['AC' + str(row_sht)] = cost_usd;
         sht['AC' + str(row_sht)].number_format = dol;

         sht['AD' + str(row_sht)] = trucks_time_to_comp_days;

         sht['AE' + str(row_sht)] = dest_time_to_comp_days;

         sht['AF' + str(row_sht)] = time_days;

         row_sht += 1;

      sum['A' + str(row_cnt)] = "ScenarioID";
      sum['A' + str(row_cnt)].font = bldu;
      sum['B' + str(row_cnt)] = scenarioid;

      sum['A' + str(row_cnt + 1)] = "Waste Type";
      sum['A' + str(row_cnt + 1)].font = bld;
      sum['B' + str(row_cnt + 1)] = sc.waste_type;

      sum['A' + str(row_cnt + 2)] = "Waste Medium";
      sum['A' + str(row_cnt + 2)].font = bld;
      sum['B' + str(row_cnt + 2)] = sc.waste_medium;

      sum['A' + str(row_cnt + 3)] = "Total Waste Amount";
      sum['A' + str(row_cnt + 3)].font = bld;
      sum['B' + str(row_cnt + 3)] = sc.waste_amount;

      sum['A' + str(row_cnt + 4)] = "Allocated Amount";
      sum['A' + str(row_cnt + 4)].font = bld;
      sum['B' + str(row_cnt + 4)] = total_allocated_amount;

      sum['A' + str(row_cnt + 5)] = "Unallocated Amount";
      sum['A' + str(row_cnt + 5)].font = bld;
      sum['B' + str(row_cnt + 5)] = unallocated_amount;

      sum['A' + str(row_cnt + 6)] = "Waste Unit";
      sum['A' + str(row_cnt + 6)].font = bld;
      sum['B' + str(row_cnt + 6)] = sc.waste_unit;
      sum['B' + str(row_cnt + 6)].alignment = rht;

      sum['C' + str(row_cnt + 1)] = "ConditionID";
      sum['C' + str(row_cnt + 1)].font = bld;
      sum['D' + str(row_cnt + 1)] = sc.conditionid;

      sum['C' + str(row_cnt + 2)] = "FactorID";
      sum['C' + str(row_cnt + 2)].font = bld;
      sum['D' + str(row_cnt + 2)] = sc.factorid;

      sum['C' + str(row_cnt + 3)] = "Total Number of Facilities";
      sum['C' + str(row_cnt + 3)].font = bld;
      sum['D' + str(row_cnt + 3)] = total_number_of_facilities;

      sum['C' + str(row_cnt + 4)] = "Total Number of Shipments";
      sum['C' + str(row_cnt + 4)].font = bld;
      sum['D' + str(row_cnt + 4)] = total_number_of_shipments;

      sum['E' + str(row_cnt + 1)] = "Total CPLM Cost ($)";
      sum['E' + str(row_cnt + 1)].font = bld;
      sum['F' + str(row_cnt + 1)] = total_cplm_cost_usd;
      sum['F' + str(row_cnt + 1)].number_format = dol;

      sum['E' + str(row_cnt + 2)] = "Total Fixed Cost ($)";
      sum['E' + str(row_cnt + 2)].font = bld;
      sum['F' + str(row_cnt + 2)] = total_fixed_cost_usd;
      sum['F' + str(row_cnt + 2)].number_format = dol;

      sum['E' + str(row_cnt + 3)] = "Total Tolls ($)";
      sum['E' + str(row_cnt + 3)].font = bld;
      sum['F' + str(row_cnt + 3)] = total_tolls_usd;
      sum['F' + str(row_cnt + 3)].number_format = dol;

      sum['E' + str(row_cnt + 4)] = "Total Misc Trans Costs ($)";
      sum['E' + str(row_cnt + 4)].font = bld;
      sum['F' + str(row_cnt + 4)] = total_misc_trans_cost_usd;
      sum['F' + str(row_cnt + 4)].number_format = dol;

      sum['E' + str(row_cnt + 5)] = "Total Transportation Cost ($)";
      sum['E' + str(row_cnt + 5)].font = bld;
      sum['F' + str(row_cnt + 5)] = total_trans_cost_usd;
      sum['F' + str(row_cnt + 5)].number_format = dol;

      sum['G' + str(row_cnt + 1)] = "Total Staging Site Cost ($)";
      sum['G' + str(row_cnt + 1)].font = bld;
      sum['H' + str(row_cnt + 1)] = total_staging_site_cost_usd;
      sum['H' + str(row_cnt + 1)].number_format = dol;

      sum['G' + str(row_cnt + 2)] = "Total Disposal Cost ($)";
      sum['G' + str(row_cnt + 2)].font = bld;
      sum['H' + str(row_cnt + 2)] = total_disposal_cost_usd;
      sum['H' + str(row_cnt + 2)].number_format = dol;

      sum['G' + str(row_cnt + 3)] = "Total Labor Cost ($)";
      sum['G' + str(row_cnt + 3)].font = bld;
      sum['H' + str(row_cnt + 3)] = total_labor_cost_usd;
      sum['H' + str(row_cnt + 3)].number_format = dol;

      sum['G' + str(row_cnt + 4)] = "Total Vehicle Decon Cost ($)";
      sum['G' + str(row_cnt + 4)].font = bld;
      sum['H' + str(row_cnt + 4)] = total_vehicle_decon_cost_usd;
      sum['H' + str(row_cnt + 4)].number_format = dol;

      sum['G' + str(row_cnt + 5)] = "Total Cost Multiplier ($)";
      sum['G' + str(row_cnt + 5)].font = bld;
      sum['H' + str(row_cnt + 5)] = total_cost_multiplier_usd;
      sum['H' + str(row_cnt + 5)].number_format = dol;
      
      sum['G' + str(row_cnt + 6)] = "Total Cost ($)";
      sum['G' + str(row_cnt + 6)].font = bld;
      sum['H' + str(row_cnt + 6)] = total_cost_usd;
      sum['H' + str(row_cnt + 6)].number_format = dol;

      sum['I' + str(row_cnt + 1)] = "Trucks Time to Complete (days)";
      sum['I' + str(row_cnt + 1)].font = bld;
      sum['J' + str(row_cnt + 1)] = max_trucks_time_to_comp_days;

      sum['I' + str(row_cnt + 2)] = "Destination Time to Complete (days)";
      sum['I' + str(row_cnt + 2)].font = bld;
      sum['J' + str(row_cnt + 2)] = max_dest_time_to_comp_days;

      sum['I' + str(row_cnt + 3)] = "Total Time Days (days)";
      sum['I' + str(row_cnt + 3)].font = bld;
      sum['J' + str(row_cnt + 3)] = max_time_days;

      row_cnt =+ 8;

   #########################################################################
   # Step 40
   # Add the Factor sheet
   #########################################################################
   ref = wb.create_sheet('Reference');

   ref.column_dimensions['A'].width = 12;
   ref.column_dimensions['B'].width = 38;
   ref.column_dimensions['C'].width = 25;
   ref.column_dimensions['D'].width = 25;
   ref.column_dimensions['E'].width = 30;
   ref.column_dimensions['F'].width = 25;
   ref.column_dimensions['G'].width = 25;
   ref.column_dimensions['H'].width = 25;

   ref['A1'] = 'Reference';
   ref['A1'].font = ft18;

   row_cnt = 3;

   for cid in ary_conditionids:
      if cid is not None:
         haz.conditions.loadConditionID(cid);

         ref['A' + str(row_cnt)] = "Condition ID";
         ref['A' + str(row_cnt)].font = bld;
         ref['B' + str(row_cnt)] = cid;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Road Tolls ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.roadtolls;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Misc Costs ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.misccost;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Total Cost Multiplier (addt\'l% of total cost)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = str(haz.conditions.totalcostmultiplier * 100) + "%";
         ref['C' + str(row_cnt)].alignment = rht;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Vehicle Decon Cost ($/shipment)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.vehicledeconcost;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Staging Site Cost ($/day)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.stagingsitecost;
         ref['C' + str(row_cnt)].number_format = dol;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Number of Trucks Available (trucks)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.numberoftrucksavailable;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = 'Driving Hours (hrs/day)';
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = haz.conditions.drivinghours;
         row_cnt += 1;

   row_cnt += 1;

   for fid in ary_factorids:
      if fid is not None:
         haz.factors.loadFactorID(fid);

         ref['A' + str(row_cnt)] = "Factor ID";
         ref['A' + str(row_cnt)].font = bld;
         ref['B' + str(row_cnt)] = fid;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Shipment Loading";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Vehicle";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Waste Type";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Waste Medium";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Loading Rate";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "Unit per shipment";
         ref['F' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.shipment_loading:
            ref['B' + str(row_cnt)] = item.vehicle;
            ref['C' + str(row_cnt)] = item.wastetype;
            ref['D' + str(row_cnt)] = item.wastemedium;
            ref['E' + str(row_cnt)] = item.loadingrate;
            ref['F' + str(row_cnt)] = item.unitpershipment;
            ref['F' + str(row_cnt)].alignment = rht;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "CPLM Unit Rates";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Vehicle";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "CPLMDist Lower";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "CPLMDist Upper";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Waste Type";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "Waste Medium";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "CPLMUnit Rate";
         ref['G' + str(row_cnt)].font = bld;
         ref['H' + str(row_cnt)] = "Unit";
         ref['H' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.cplm_unit_rates:
            ref['B' + str(row_cnt)] = item.vehicle;
            ref['C' + str(row_cnt)] = item.cplmdist_lower;
            ref['D' + str(row_cnt)] = item.cplmdist_upper;
            ref['E' + str(row_cnt)] = item.wastetype;
            ref['F' + str(row_cnt)] = item.wastemedium;
            ref['G' + str(row_cnt)] = item.cplunit_rate;
            ref['H' + str(row_cnt)] = item.unit;
            ref['H' + str(row_cnt)].alignment = rht;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Fixed Trans Cost";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Vehicle";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "FixedCost Type";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Waste Type";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Waste Medium";
         ref['E' + str(row_cnt)].font = bld;
         ref['F' + str(row_cnt)] = "FixedCost Value";
         ref['F' + str(row_cnt)].font = bld;
         ref['G' + str(row_cnt)] = "Unit";
         ref['G' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.fixed_trans_cost:
            ref['B' + str(row_cnt)] = item.vehicle;
            ref['C' + str(row_cnt)] = item.fixedcost_type;
            ref['D' + str(row_cnt)] = item.wastetype;
            ref['E' + str(row_cnt)] = item.wastemedium;
            ref['F' + str(row_cnt)] = item.fixedcost_value;
            ref['G' + str(row_cnt)] = item.unit;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Labor Costs";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Labor Category";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Labor Cost";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Unit";
         ref['D' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.labor_costs:
            ref['B' + str(row_cnt)] = item.laborcategory;
            ref['C' + str(row_cnt)] = item.laborcost;
            ref['D' + str(row_cnt)] = item.unit;
            row_cnt += 1;

         ref['B' + str(row_cnt)] = "Disposal Fees";
         ref['B' + str(row_cnt)].font = bld;
         row_cnt += 1;

         ref['B' + str(row_cnt)] = "Waste Type";
         ref['B' + str(row_cnt)].font = bld;
         ref['C' + str(row_cnt)] = "Waste Medium";
         ref['C' + str(row_cnt)].font = bld;
         ref['D' + str(row_cnt)] = "Disposal Cost";
         ref['D' + str(row_cnt)].font = bld;
         ref['E' + str(row_cnt)] = "Unit";
         ref['E' + str(row_cnt)].font = bld;
         row_cnt += 1;

         for item in haz.factors.disposal_fees:
            ref['B' + str(row_cnt)] = item.wastetype;
            ref['C' + str(row_cnt)] = item.wastemedium;
            ref['D' + str(row_cnt)] = item.disposalcost;
            ref['E' + str(row_cnt)] = item.unit;
            row_cnt += 1;

         row_cnt += 1;

   #########################################################################
   # Step 50
   # Write out the excel file
   #########################################################################
   wb.save(dest_filename);

   #########################################################################
   # Step 60
   #
   #########################################################################
   del wb,sum,sht;
   del cursor_in;
   del haz,sc;

   return;
